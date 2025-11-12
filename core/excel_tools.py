from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelTemplateHandler:
    def __init__(self, template_path, output_path):
        self.wb = load_workbook(template_path)
        self.output_path = output_path
        
    def fill_data(self, sheet_data):
        for sheet_name, data in sheet_data.items():
            ws = self.wb[sheet_name]
            for region, values in data.items():
                start_cell = region.split(':')[0]
                for i, row in enumerate(values):
                    for j, value in enumerate(row):
                        cell = ws[f"{get_column_letter(j+1)}{i+start_cell}"]
                        cell.value = value
    
    def preserve_formulas(self, formula_cells):
        for cell_ref in formula_cells:
            sheet, cell = cell_ref.split('!')
            self.wb[sheet][cell].value = self.wb[sheet][cell].value
    
    def save(self):
        self.wb.save(self.output_path)